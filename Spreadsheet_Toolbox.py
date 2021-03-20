#!/usr/bin/env python3
# Ricardo Cebreros Campos
"""
This module contains the tools necessary to work with spreadsheets.
Complete break-down on each function can be found on the README.md file.
"""
import openpyxl, os
from datetime import datetime

# Created a list with all files in a given directory.
# It then iterates through each file to check for CSV or XLSX files.
def get_folder_records(pathName):
    all_rows = []
    files = os.listdir(pathName)
    for file in files:
        # os.path.sep returns '\' or '/' depending on the OS used.
        path = pathName + os.path.sep + file
        print(path)
        if '.csv' == file[-4:].lower():
            content = parse_csv(path)
            all_rows.append(content)
        elif '.xlsx' == file[-5:].lower():
            content = parse_xlsx(path)
            all_rows.append(content)
    return all_rows

# Returns a 2 dimensional list with all contents of a spreadsheet.
# Each row is added as a list.
def parse_xlsx(file):
    contents = []
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.active
    for line in worksheet.values:
        line = list(line)
        while len(line) >= 1 and line[-1] == None:
            line.pop()
        if len(line) > 0:
            contents.append(line)
    workbook.close()
    return contents

# Returns a 2 dimensional list with all contents of a spreadsheet.
# Each row is added as a list.
def parse_csv(file):
    contents = []
    with open(file, 'r') as record:
        for line in record.readlines():
            line = line.replace('\n', '')
            contents.append(line.split(','))
    return contents

# Returns dictionary.
# The dictionary will have a key and the entire row will be the value.
# FILE - Directory path.
# Key-Num - index that will become the designated key.
# Include Shorts - Will include rows where the row is shorter than the given key index.
# Append_Repeats - If a key repeats itself, it will append the row to the existing value.
def sheet_to_dict(file, key_num, include_shorts = False, append_repeats = True):
    content = {'shortlines': []}
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.active
    for line in worksheet.values:
        line = list(line)
        while len(line) >= 1 and line[-1] == None:
            line.pop()

        if len(line) > key_num and key_num not in content.keys():
            content[line[key_num]] = line
            continue

        elif key_num in content.keys() and append_repeats == True:
            key_num[key_num].append(line)
            continue

        elif len(line) > 0:
            content['shortlines'].append(line)

    if include_shorts == False:
        del content['shortlines']
    workbook.close()
    return content


# Returns a specified column in a given spreadsheet.
def get_column(file, index):
    column = []

    if file[-5:] == '.xlsx':
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook.active
        for line in worksheet.values:
            line = list(line)
            if len(line) > index:
                column.append(line[index])
        workbook.close()

    if file[-4:] == '.csv':
        with open(file, 'r') as filename:
            for line in filename.readlines():
                line = line.split(',')
                if len(line) > index:
                    column.append(line[index].replace('\n', ''))
    return column


# Takes a 2 dimensional list and writes the lists into rows.
# PASRSE can be set to CSV if the lists are made up of CSV rows.
def save_results(data= None, title= None, parse=None):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if parse == 'csv':
        for line in data:
            worksheet.append(line.split(','))
    else:
        worksheet.append(data)
    workbook.save(title)


# Returns a datetime object.
def date_object(date):
    if len(date) > 10:
        # Excludes time entries, takes only date on first half.
        date = date.split(' ')[0]
    return datetime.strptime(date.strip(), '%m/%d/%Y').date()

# Returns a string from a given datetime object.
def string_date(date):
    return datetime.strftime(date, '%m/%d/%Y')